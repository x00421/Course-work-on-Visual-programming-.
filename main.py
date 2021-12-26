from PyQt5 import QtWidgets
from PyQt5 import QtCore, QtGui
from openpyxl.worksheet import worksheet
from main_ui import Ui_Main  # импорт нашего сгенерированного файла

import datetime
from katalog import showkatalog

from AddNewPosition import showAddNewPosition
from vozvrat import showvozvrat
from otchet import showOtchet

import sys
import openpyxl
import time
import pandas as pd
pol_global = "МЖ"


class mywindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(mywindow, self).__init__()
        self.ui = Ui_Main()
        self.ui.setupUi(self)

        # Отвечает за выбор размера в поиске
        self.ui.SizecomboBox.addItem("*")
        self.ui.SizecomboBox.addItem("XS")
        self.ui.SizecomboBox.addItem("S")
        self.ui.SizecomboBox.addItem("M")
        self.ui.SizecomboBox.addItem("L")
        self.ui.SizecomboBox.addItem("XL")
        self.ui.SizecomboBox.addItem("XXL")

        # Отвечает за Отображение колонок
        self.ui.MaintableWidget.setColumnCount(4)
        self.ui.MaintableWidget.setColumnWidth(0, 450)
        self.ui.MaintableWidget.setColumnWidth(1, 60)
        self.ui.MaintableWidget.setColumnWidth(2, 130)
        self.ui.MaintableWidget.setColumnWidth(3, 100)
        self.ui.MaintableWidget.setHorizontalHeaderLabels(
            ('Название', 'Размер', 'Цена', 'Остаток')
        )
        # Функция отображения таблицы без фильтра
        self.mainloaddata()
        self.ui.AllButton.clicked.connect(self.mainloaddata)
        # Функция отображения таблицы для мужчин
        self.ui.MenButton.clicked.connect(self.mainloaddata_man)
        # Функция отображения таблицы для женщин
        self.ui.WomenButton.clicked.connect(self.mainloaddata_women)
        # Функция поиска
        self.ui.PoiskButton.clicked.connect(self.poisk)
        
        self.ui.OtpriceTextEdit.setValidator(QtGui.QIntValidator())
        self.ui.DopriceTextEdit.setValidator(QtGui.QIntValidator())

        self.ui.DelButto.clicked.connect(self.delete)

        
        self.ui.AddButton.clicked.connect(lambda: showAddNewPosition(self))
        self.ui.ProdagaButton.clicked.connect(self.prodaga)
        self.ui.KatalogButton.clicked.connect(lambda: showkatalog(self))
        self.ui.TovarButton_2.clicked.connect(lambda: showvozvrat(self))
        self.ui.OtchetButton.clicked.connect(lambda:showOtchet(self))
        self.ui.MaintableWidget.clicked.connect(self. vizual)

   

    def mainloaddata(self):  # Функция отображения таблицы без фильтра
        
        global  pol_global
        pol_global="МЖ"
        
        book = openpyxl.open("baza.xlsx")
        sheet = book.worksheets[0]

        row = 0
        self.ui.MaintableWidget.setRowCount(sheet.max_row-1)
        for i in range( sheet.max_row):
            self.ui.MaintableWidget.showRow(i)

        for i in range(2, sheet.max_row+1):
            
            name = sheet['A'+str(i)].value
            size = sheet["C"+str(i)].value
            price = sheet["F"+str(i)].value
            number = sheet["G"+str(i)].value
            self.ui.MaintableWidget.setItem(
                row, 0, QtWidgets.QTableWidgetItem(str(name)))
            self.ui.MaintableWidget.setItem(
                row, 1, QtWidgets.QTableWidgetItem(str(size)))
            self.ui.MaintableWidget.setItem(
                row, 2, QtWidgets.QTableWidgetItem(str(price)))
            self.ui.MaintableWidget.setItem(
                row, 3, QtWidgets.QTableWidgetItem(str(number)))
            row += 1

    def mainloaddata_man(self):  # Функция отображения таблицы для мужчин
        global  pol_global
        pol_global="М"
        book = openpyxl.open("baza.xlsx")
        sheet = book.worksheets[0]
        for i in range( sheet.max_row):
            self.ui.MaintableWidget.showRow(i)
        

        row = 0
        for i in range(2, sheet.max_row+1):
            pol = sheet['J'+str(i)].value
            if pol!="М":
                self.ui.MaintableWidget.hideRow(row)
            row = row+1

    def mainloaddata_women(self):  # Функция отображения таблицы для Женщин
        global  pol_global
        pol_global="Ж"
        book = openpyxl.open("baza.xlsx")
        sheet = book.worksheets[0]
        for i in range( sheet.max_row):
            self.ui.MaintableWidget.showRow(i)
        row = 0
        for i in range(2, sheet.max_row+1):
            pol = sheet['J'+str(i)].value
            if pol!="Ж":
                self.ui.MaintableWidget.hideRow(row)
            row = row+1

    def poisk(self):
        name_p = self.ui.NameEdit.toPlainText()  # Получаем текст из строкиввода
        size_p = self.ui.SizecomboBox.currentText()  # Получаем текст из комбобокса
        minprice = self.ui.OtpriceTextEdit.text()
        maxprice = self.ui.DopriceTextEdit.text()
        global  pol_global
       

        if name_p == "":
            msg = QtWidgets.QMessageBox()
            msg.setWindowTitle("Ошибка!")
            msg.setText("Введите название!")
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.exec_()
        else:
            if size_p == "*" and minprice == "" and maxprice == "":
                msg = QtWidgets.QMessageBox()
                msg.setWindowTitle("Ошибка!")
                msg.setText("Введите стоимость или размер!")
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.exec_()

            else:
                book = openpyxl.open("baza.xlsx")
                sheet = book.worksheets[0]
                row = 0
                self.ui.MaintableWidget.setRowCount(sheet.max_row-1)
                for i in range( sheet.max_row):
                    self.ui.MaintableWidget.hideRow(i)
               
                self.ui.MaintableWidget.setHorizontalHeaderLabels(
                    ('Название', 'Размер', 'Цена', 'Остаток')
                )
                for i in range(2, sheet.max_row):
                    name = self.ui.MaintableWidget.item(i-2,0).text()
                    size = self.ui.MaintableWidget.item(i-2,1).text()
                    price = self.ui.MaintableWidget.item(i-2,2).text()
                    number = self.ui.MaintableWidget.item(i-2,3).text()
                    pol = sheet['J'+str(i)].value
                    print(name)
                    

                    if size_p == "*":#Если пустое поле размера
                            if maxprice == "":
                                maxprice = 100000000
                            if minprice == "":
                                minprice = 0
                            if (name_p in name)and(pol in pol_global) and ((int(minprice) <= int(price)) and (int(maxprice) >= int(price))):
                                self.ui.MaintableWidget.showRow(i-2)
                                row = row+1

                    else:#Если поле размера не пустое 
                        if maxprice == "" and minprice == "":#Если поле размера не пустое а цены пустые
                            if (name_p in name) and size == size_p and (pol in pol_global):
                                self.ui.MaintableWidget.showRow(i-2)
                                row = row+1
                        else:
                            if maxprice == "":
                                maxprice = 100000000
                            if minprice == "":
                                minprice = 0
                            if (name_p in name)  and (pol in pol_global) and (size == size_p) and ((int(minprice) <= int(price)) and (int(maxprice) >= int(price))):
                                self.ui.MaintableWidget.showRow(i-2)
                                row = row+1

    def getData(self):#Функция получения информации из таблицы
        rows = self.ui.MaintableWidget.rowCount()
        cols = self.ui.MaintableWidget.columnCount()
        data = []
        flag=0
        for row in range(rows):
            tmp = []
            if flag==1:
                break
            for col in range(cols):
                if col==0 and self.ui.MaintableWidget.item(row,0).text()=="":
                    flag=1
                    break 
                try:
                    tmp.append(self.ui.MaintableWidget.item(row,col).text())
                except:
                    tmp.append('No data')
            data.append(tmp)
        for i in data: print(i)

    def delete(self):
       
        row=self.ui.MaintableWidget.currentRow() #Получение нужной строки
        
        if row ==-1:
            msg = QtWidgets.QMessageBox()
            msg.setWindowTitle("Ошибка!")
            msg.setText("Выберите строку!")
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.exec_()
        else:
            filename="baza.xlsx"
            book = openpyxl.load_workbook(filename=filename)
            sheet :worksheet= book.worksheets[0]
            sheet.delete_rows(row+2)
            book.save(filename)
            print(row)
            self.mainloaddata()

    def prodaga(self):
        row_prodaga=self.ui.MaintableWidget.currentRow() 
        print(row_prodaga)
        if row_prodaga ==-1:
            msg = QtWidgets.QMessageBox()
            msg.setWindowTitle("Ошибка!")
            msg.setText("Выберите строку!")
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.exec_()
        else:
            filename="baza.xlsx"
            book = openpyxl.load_workbook(filename=filename)
            sheet :worksheet= book.worksheets[0]
            name = sheet['A'+str(row_prodaga+2)].value
            cod=sheet['B'+str(row_prodaga+2)].value
            size = sheet["C"+str(row_prodaga+2)].value
            proizv=sheet['D'+str(row_prodaga+2)].value
            postav=sheet['E'+str(row_prodaga+2)].value
            price = sheet["F"+str(row_prodaga+2)].value
            number = sheet["G"+str(row_prodaga+2)].value
            date=datetime.datetime.now()
            date.strftime("%Y,%m,%d")
            photo=sheet['I'+str(row_prodaga+2)].value
            pol=sheet['J'+str(row_prodaga+2)].value
            if int(number)>0:
                sheet["G"+str((row_prodaga+2))].value=str(int(sheet["G"+str((row_prodaga+2))].value)-1)
                number=1
                sheet :worksheet= book.worksheets[1]
                sheet.insert_rows(2)
                sheet["A2"].value=str(name)
                sheet["B2"].value=str(cod)
                sheet["C2"].value=str(size)
                sheet["D2"].value=str(proizv)
                sheet["E2"].value=str(postav)
                sheet["F2"].value=str(price)
                sheet["G2"].value=str(number)
                sheet["H2"].value=date.strftime("%Y,%m,%d")
                sheet["I2"].value=str(photo)
                sheet["J2"].value=str(pol)
                sheet["K2"].value=str("П")

            else:
                msg = QtWidgets.QMessageBox()
                msg.setWindowTitle("Ошибка!")
                msg.setText("Товар закончился")
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.exec_()



            book.save(filename)
            
            self.mainloaddata()

    def vizual(self):
        row=self.ui.MaintableWidget.currentRow()
        i=int(self.ui.MaintableWidget.item(row,3).text())
        if i>15:
            i=15
        self.ui.progressBar.setValue(i)

        print(0)

    

                       


app = QtWidgets.QApplication([])
application = mywindow()
application.show()

sys.exit(app.exec())
