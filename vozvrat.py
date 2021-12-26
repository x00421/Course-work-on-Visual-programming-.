from PyQt5 import QtWidgets
from vozvrat_ui import Ui_Form
import sys
from PyQt5.QtGui import QPixmap
import openpyxl
from openpyxl.worksheet import worksheet


class Vozvratwindow(QtWidgets.QMainWindow):
        def __init__(self, parent=None):
                super(Vozvratwindow, self).__init__(parent)
                self.ui = Ui_Form()
                self.ui.setupUi(self)
                self.ui.MaintableWidget.setColumnCount(5)
                self.ui.MaintableWidget.setColumnWidth(0, 250)
                self.ui.MaintableWidget.setColumnWidth(1, 60)
                self.ui.MaintableWidget.setColumnWidth(2, 130)
                self.ui.MaintableWidget.setColumnWidth(3, 60)
                self.ui.MaintableWidget.setColumnWidth(4, 140)
                self.ui.MaintableWidget.setHorizontalHeaderLabels(
                    ('Название', 'Размер', 'Цена', 'Статус','Дата')
                )
                self.mainloaddata()
                self.ui.VozvratButton.clicked.connect(self.vozvrat)
                self.ui.BrakButton.clicked.connect(self.brak)
                self.ui.SortButton.clicked.connect(self.sorttable)

        
        
        def mainloaddata(self):
            book = openpyxl.open("baza.xlsx")
            sheet = book.worksheets[1]
            for i in range( sheet.max_row):
                self.ui.MaintableWidget.showRow(i)
            row = 0
            self.ui.MaintableWidget.setRowCount(sheet.max_row-1)
            for i in range(2, sheet.max_row+1):
                name = sheet['A'+str(i)].value
                size = sheet["C"+str(i)].value
                price = sheet["F"+str(i)].value
                number = sheet["K"+str(i)].value
                date=sheet["H"+str(i)].value
                self.ui.MaintableWidget.setItem(
                    row, 0, QtWidgets.QTableWidgetItem(str(name)))
                self.ui.MaintableWidget.setItem(
                    row, 1, QtWidgets.QTableWidgetItem(str(size)))
                self.ui.MaintableWidget.setItem(
                    row, 2, QtWidgets.QTableWidgetItem(str(price)))
                self.ui.MaintableWidget.setItem(
                    row, 3, QtWidgets.QTableWidgetItem(str(number)))
                self.ui.MaintableWidget.setItem(
                    row, 4, QtWidgets.QTableWidgetItem(date))
                row += 1
        
        def vozvrat(self):
            row_vozvrat=self.ui.MaintableWidget.currentRow()      
            filename="baza.xlsx"
            book = openpyxl.load_workbook(filename=filename)
            sheet :worksheet= book.worksheets[1]
            sheet["K"+str((row_vozvrat+2))].value=str("В")
            sheet["F"+str((row_vozvrat+2))].value=str("0")
            #Вернём наличие товара в базу
            cod_vozvrata=sheet["B"+str((row_vozvrat+2))].value
            print(cod_vozvrata, "-cod_vozvrata")
            sheet1 :worksheet= book.worksheets[0]
            for i in range(2,sheet.max_row):
                if sheet1["B"+str((row_vozvrat+2))].value==cod_vozvrata:
                    nalich=int(sheet1["G"+str((row_vozvrat+2))].value)
                    sheet1["G"+str((row_vozvrat+2))].value=str(nalich+1)
                    print(sheet1["G"+str((row_vozvrat+2))].value)
                    book.save(filename)
                    self.mainloaddata()
                    return
            
               

        def brak(self):
            row_brak=self.ui.MaintableWidget.currentRow()
            print(row_brak)  
            filename="baza.xlsx"
            book = openpyxl.load_workbook(filename=filename)
            sheet :worksheet= book.worksheets[1]
            sheet["K"+str((row_brak+2))].value=str("Б")
            sheet["F"+str((row_brak+2))].value=str("0")
            book.save(filename)
            self.mainloaddata()
               
        def sorttable(self):
            filename="baza.xlsx"
            book = openpyxl.load_workbook(filename=filename)
            sheet :worksheet= book.worksheets[1]

            for i in range( sheet.max_row):
                    self.ui.MaintableWidget.hideRow(i)

            for i in range(2, sheet.max_row):
                
                    name = self.ui.MaintableWidget.item(i-2,0).text()
                    size = self.ui.MaintableWidget.item(i-2,1).text()
                    price = self.ui.MaintableWidget.item(i-2,2).text()
                    status = self.ui.MaintableWidget.item(i-2,3).text()
                    date = self.ui.MaintableWidget.item(i-2,4).text()

                    dateot=self.ui.otdateEdit.date().toString("yyyy,M,d")
                    datedo=self.ui.dateEdit_2.date().toString("yyyy,M,d")
                    print(date)
                    if dateot<=date and datedo>=date:
                         self.ui.MaintableWidget.showRow(i-2)





def showvozvrat(app):
    application = Vozvratwindow(app)
    application.show()